import os
import joblib
import pandas as pd
import numpy as np
from tabulate import tabulate
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Model performanslarını karşılaştıran ve analiz eden modül.
def load_performance_metrics(model_type, interval, metric_type='r2'):
    # Belirli bir model türü için performans metriklerini yükler.
    base_dir = "d:/Crypto_Analysis_random_state/models"
    model_dir = os.path.join(base_dir, model_type, interval)
    
    if not os.path.exists(model_dir):
        return None
    
    metrics = []
    for file in os.listdir(model_dir):
        if file.endswith("_performance.joblib"):
            performance = joblib.load(os.path.join(model_dir, file))
            
            # Metrik tipine göre sütun adını belirle
            if metric_type == 'r2':
                column_suffix = 'R²'
                metric_key = 'r2_test'
            elif metric_type == 'mae':
                column_suffix = 'MAE'
                metric_key = 'mae_test'
            elif metric_type == 'rmse':
                column_suffix = 'RMSE'
                metric_key = 'rmse_test'
            else:
                raise ValueError(f"Bilinmeyen metrik tipi: {metric_type}")
            
            column_name = {
                'gradient_boosting': f'Test {column_suffix} (GB)',
                'random_forest': f'Test {column_suffix} (RF)',
                'decision_tree': f'Test {column_suffix} (DT)',
                'svr': f'Test {column_suffix} (SVR)',
                'ridge_regression': f'Test {column_suffix} (Ridge)',
                'lightgbm': f'Test {column_suffix} (LGBM)',
                'knn': f'Test {column_suffix} (KNN)'
            }[model_type]
            
            # Eğer metrik yoksa, varsayılan değer kullan
            metric_value = performance.get(metric_key, np.nan)
            
            metrics.append({
                'symbol': performance['symbol'],
                column_name: metric_value
            })
    
    if not metrics:
        return None
        
    return pd.DataFrame(metrics)

def print_model_comparison(gb_metrics, rf_metrics, dt_metrics, svr_metrics, ridge_metrics, lgbm_metrics, knn_metrics):
    # Modellerin performansını karşılaştırır ve en iyisini belirler.
    if gb_metrics is None or rf_metrics is None or dt_metrics is None or svr_metrics is None or ridge_metrics is None or lgbm_metrics is None or knn_metrics is None:
        return
    
    # Ortalama hesaplama
    gb_test = gb_metrics['Test R² (GB)'].mean()
    rf_test = rf_metrics['Test R² (RF)'].mean()
    dt_test = dt_metrics['Test R² (DT)'].mean()
    svr_test = svr_metrics['Test R² (SVR)'].mean()
    ridge_test = ridge_metrics['Test R² (Ridge)'].mean()
    lgbm_test = lgbm_metrics['Test R² (LGBM)'].mean()
    knn_test = knn_metrics['Test R² (KNN)'].mean()
    
    # En iyi model
    scores = {
        'GB': gb_test,
        'RF': rf_test,
        'DT': dt_test,
        'SVR': svr_test,
        'Ridge': ridge_test,
        'LGBM': lgbm_test,
        'KNN': knn_test
    }
    winner = max(scores.items(), key=lambda x: x[1])[0]
    
    # Sonuçlar
    # print("\nModel Karşılaştırmaları:")
    # print("-" * 50)
    # print(f"GB - Ortalama Test R²: {gb_test:.4f}")
    # print(f"RF - Ortalama Test R²: {rf_test:.4f}")
    # print(f"DT - Ortalama Test R²: {dt_test:.4f}")
    # print(f"SVR - Ortalama Test R²: {svr_test:.4f}")
    # print(f"Ridge - Ortalama Test R²: {ridge_test:.4f}")
    # print(f"LGBM - Ortalama Test R²: {lgbm_test:.4f}")
    # print(f"KNN - Ortalama Test R²: {knn_test:.4f}")
    # print(f"En İyi Model: {winner}")

def create_comparison_sheet(all_metrics, metric_type='r2'):
    # Tüm zaman aralıkları için karşılaştırma özeti oluşturur.
    
    comparison_data = []
    
    # Metrik tipine göre sütun adlarını belirle
    if metric_type == 'r2':
        metric_suffix = 'R²'
        title_suffix = 'R²'
    elif metric_type == 'mae':
        metric_suffix = 'MAE'
        title_suffix = 'MAE'
    elif metric_type == 'rmse':
        metric_suffix = 'RMSE'
        title_suffix = 'RMSE'
    else:
        raise ValueError(f"Bilinmeyen metrik tipi: {metric_type}")
    
    # Her zaman aralığı için ortalama performansları hesapla
    for interval, metrics in all_metrics.items():
        avg_performance = {
            'Zaman Aralığı': interval,
            f'GB Ortalama {metric_suffix}': metrics[f'Test {metric_suffix} (GB)'].mean(),
            f'RF Ortalama {metric_suffix}': metrics[f'Test {metric_suffix} (RF)'].mean(),
            f'DT Ortalama {metric_suffix}': metrics[f'Test {metric_suffix} (DT)'].mean(),
            f'SVR Ortalama {metric_suffix}': metrics[f'Test {metric_suffix} (SVR)'].mean(),
            f'Ridge Ortalama {metric_suffix}': metrics[f'Test {metric_suffix} (Ridge)'].mean(),
            f'LGBM Ortalama {metric_suffix}': metrics[f'Test {metric_suffix} (LGBM)'].mean(),
            f'KNN Ortalama {metric_suffix}': metrics[f'Test {metric_suffix} (KNN)'].mean()
        }
        
        # En iyi model ve skoru (R² için en yüksek, MAE/RMSE için en düşük)
        models = {
            'GB': avg_performance[f'GB Ortalama {metric_suffix}'],
            'RF': avg_performance[f'RF Ortalama {metric_suffix}'],
            'DT': avg_performance[f'DT Ortalama {metric_suffix}'],
            'SVR': avg_performance[f'SVR Ortalama {metric_suffix}'],
            'Ridge': avg_performance[f'Ridge Ortalama {metric_suffix}'],
            'LGBM': avg_performance[f'LGBM Ortalama {metric_suffix}'],
            'KNN': avg_performance[f'KNN Ortalama {metric_suffix}']
        }
        
        # R² için en yüksek, MAE/RMSE için en düşük değer en iyidir
        if metric_type == 'r2':
            best_model = max(models.items(), key=lambda x: x[1])
        else:
            best_model = min(models.items(), key=lambda x: x[1])
            
        avg_performance['En İyi Model'] = best_model[0]
        avg_performance['En İyi Skor'] = best_model[1]
        
        comparison_data.append(avg_performance)
    
    # DataFrame oluştur
    comparison_df = pd.DataFrame(comparison_data)
    
    # Sütun sıralamasını düzenle
    column_order = [
        'Zaman Aralığı',
        'En İyi Model',
        'En İyi Skor',
        f'GB Ortalama {metric_suffix}',
        f'RF Ortalama {metric_suffix}',
        f'DT Ortalama {metric_suffix}',
        f'SVR Ortalama {metric_suffix}',
        f'Ridge Ortalama {metric_suffix}',
        f'LGBM Ortalama {metric_suffix}',
        f'KNN Ortalama {metric_suffix}'
    ]
    comparison_df = comparison_df[column_order]
    
    # Sayısal değerleri formatla
    for col in comparison_df.columns:
        if col not in ['Zaman Aralığı', 'En İyi Model']:
            # R² için yüzde formatı, MAE/RMSE için ondalık formatı kullan
            if metric_type == 'r2':
                comparison_df[col] = comparison_df[col].apply(lambda x: f"{x:.2%}")
            else:
                comparison_df[col] = comparison_df[col].apply(lambda x: f"{x:.4f}")
    
    return comparison_df

def apply_excel_styling(worksheet, df):
    # Excel sayfasına stil uygular.
    # Renk tanımları
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    best_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    model_header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    
    # Kenarlık stilleri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlık stili
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    model_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    
    # Model başlıklarını ekle
    worksheet.insert_rows(1)
    model_headers = {
        'Zaman Aralığı': ('', 1),
        'En İyi Model': ('En İyi Model', 2),
        'GB Ortalama R²': ('Gradient Boosting', 1),
        'RF Ortalama R²': ('Random Forest', 1),
        'DT Ortalama R²': ('Decision Tree', 1),
        'SVR Ortalama R²': ('Support Vector Regression', 1),
        'Ridge Ortalama R²': ('Ridge Regression', 1),
        'LGBM Ortalama R²': ('LightGBM', 1),
        'KNN Ortalama R²': ('K-Nearest Neighbors', 1)
    }
    
    current_col = 1
    for col_name in df.columns:
        if col_name in model_headers:
            header_text, span = model_headers[col_name]
            if header_text:
                cell = worksheet.cell(row=1, column=current_col)
                cell.value = header_text
                cell.font = model_font
                cell.fill = model_header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                if span > 1:
                    worksheet.merge_cells(
                        start_row=1,
                        start_column=current_col,
                        end_row=1,
                        end_column=current_col + span - 1
                    )
        current_col += 1
    
    # Sütun genişliklerini ayarla
    for col in range(len(df.columns)):
        column_letter = get_column_letter(col + 1)
        max_length = 0
        column = df.iloc[:, col]
        
        # Başlık uzunluğunu kontrol et
        header_length = len(str(df.columns[col]))
        # Model başlığı uzunluğunu da kontrol et
        if df.columns[col] in model_headers:
            header_length = max(header_length, len(model_headers[df.columns[col]][0]))
        max_length = max(max_length, header_length)
        
        # Hücre içeriği uzunluklarını kontrol et
        for cell in column:
            try:
                if len(str(cell)) > max_length:
                    max_length = len(str(cell))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Alt başlıkları uygula
    for col in range(len(df.columns)):
        cell = worksheet.cell(row=2, column=col + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Veri hücrelerini formatla
    for row in range(3, len(df) + 3):
        # Alternatif satır renklendirmesi
        row_fill = alt_row_fill if row % 2 == 0 else PatternFill()
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = row_fill
            
            # En iyi skoru vurgula
            if df.columns[col-1] == 'En İyi Skor':
                cell.fill = best_score_fill
    
    # Başlık satırını dondur
    worksheet.freeze_panes = "A3"

def export_to_excel(all_metrics_r2, all_metrics_mae, all_metrics_rmse):
    # Tüm metrikleri tek bir Excel dosyasına aktarır.
    
    # Sonuçlar klasörünü oluştur
    output_dir = "sonuçlar"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Dosya adını oluştur
    filename = f"{output_dir}/sonuclar.xlsx"
    
    # Eğer dosya varsa, üzerine yaz
    if os.path.exists(filename):
        try:
            os.remove(filename)
        except PermissionError:
            print(f"Hata: {filename} dosyası başka bir uygulama tarafından kullanılıyor. Lütfen dosyayı kapatın ve tekrar deneyin.")
            return
    
    # Excel yazıcı oluştur
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Her zaman aralığı için ayrı sayfa oluştur
        for interval in all_metrics_r2.keys():
            # Modeller ve metrik tipleri
            models = ['GB', 'RF', 'DT', 'SVR', 'Ridge', 'LGBM', 'KNN']
            metric_types = ['R²', 'MAE', 'RMSE']
            
            # Kripto para sembollerini al
            symbols = all_metrics_r2[interval]['symbol'].tolist()
            
            # Sonuç DataFrame'leri oluştur
            r2_df = pd.DataFrame({'Kripto Para': symbols})
            mae_df = pd.DataFrame({'Kripto Para': symbols})
            rmse_df = pd.DataFrame({'Kripto Para': symbols})
            
            # Her model için metrikleri ekle
            for model in models:
                # R² değerlerini al
                r2_col = f'Test R² ({model})'
                if r2_col in all_metrics_r2[interval].columns:
                    r2_df[f'{model}'] = all_metrics_r2[interval][r2_col]
                
                # MAE değerlerini al
                mae_col = f'Test MAE ({model})'
                if mae_col in all_metrics_mae[interval].columns:
                    mae_df[f'{model}'] = all_metrics_mae[interval][mae_col]
                
                # RMSE değerlerini al
                rmse_col = f'Test RMSE ({model})'
                if rmse_col in all_metrics_rmse[interval].columns:
                    rmse_df[f'{model}'] = all_metrics_rmse[interval][rmse_col]
            
            # Tüm DataFrame'leri birleştir (alt alta)
            # Başlık satırları ekle
            r2_header = pd.DataFrame({'Kripto Para': ['R² Metrics']})
            mae_header = pd.DataFrame({'Kripto Para': ['MAE Metrics']})
            rmse_header = pd.DataFrame({'Kripto Para': ['RMSE Metrics']})
            
            # Model sütunlarını ekle
            for model in models:
                r2_header[model] = ['']
                mae_header[model] = ['']
                rmse_header[model] = ['']
            
            # DataFrameleri birleştir
            combined_df = pd.concat([
                r2_header,
                r2_df,
                pd.DataFrame({'Kripto Para': ['']}),  # Boş satır
                mae_header,
                mae_df,
                pd.DataFrame({'Kripto Para': ['']}),  # Boş satır
                rmse_header,
                rmse_df
            ], ignore_index=True)
            
            # Excel'e yaz
            combined_df.to_excel(writer, sheet_name=f"{interval}", index=False)
            
            # Sayfaya stil uygula
            worksheet = writer.sheets[f"{interval}"]
            
            # Başlık satırları için stil
            header_rows = [0, len(r2_df) + 2, len(r2_df) + len(mae_df) + 4]
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            metric_header_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
            
            # Metrik başlıkları için stil
            for row_idx in header_rows:
                for col_idx in range(1, len(models) + 2):
                    cell = worksheet.cell(row=row_idx + 1, column=col_idx)
                    cell.font = Font(name='Calibri', size=12, bold=True, color="000000")
                    cell.fill = metric_header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Sütun başlıkları için stil
            for col_idx in range(1, len(models) + 2):
                # R² başlıkları
                cell = worksheet.cell(row=2, column=col_idx)
                cell.font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # MAE başlıkları
                cell = worksheet.cell(row=len(r2_df) + 4, column=col_idx)
                cell.font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # RMSE başlıkları
                cell = worksheet.cell(row=len(r2_df) + len(mae_df) + 6, column=col_idx)
                cell.font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Veri hücreleri için stil
            # R² bölümü
            for row_idx in range(len(r2_df)):
                row_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid") if row_idx % 2 == 0 else PatternFill()
                for col_idx in range(1, len(models) + 2):
                    cell = worksheet.cell(row=row_idx + 3, column=col_idx)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = row_fill
                    
                    # Sayısal değerler için format
                    if col_idx > 1:  # İlk sütun kripto adı
                        cell.number_format = '0.00%'
            
            # MAE bölümü
            for row_idx in range(len(mae_df)):
                row_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid") if row_idx % 2 == 0 else PatternFill()
                for col_idx in range(1, len(models) + 2):
                    cell = worksheet.cell(row=row_idx + len(r2_df) + 5, column=col_idx)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = row_fill
                    
                    # Sayısal değerler için format
                    if col_idx > 1:  # İlk sütun kripto adı
                        cell.number_format = '0.0000'
            
            # RMSE bölümü
            for row_idx in range(len(rmse_df)):
                row_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid") if row_idx % 2 == 0 else PatternFill()
                for col_idx in range(1, len(models) + 2):
                    cell = worksheet.cell(row=row_idx + len(r2_df) + len(mae_df) + 7, column=col_idx)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = row_fill
                    
                    # Sayısal değerler için format
                    if col_idx > 1:  # İlk sütun kripto adı
                        cell.number_format = '0.0000'
            
            # Sütun genişliklerini ayarla
            for col in range(len(models) + 1):
                column_letter = get_column_letter(col + 1)
                if col == 0:  # Kripto Para sütunu
                    worksheet.column_dimensions[column_letter].width = 15
                else:  # Model sütunları
                    worksheet.column_dimensions[column_letter].width = 12
            
            # Başlık satırını dondur
            worksheet.freeze_panes = "A3"
    
    print(f"sonuçları Excel dosyasına kaydedildi.")

def display_metrics():
    # Tüm modellerin performans metriklerini yükler ve karşılaştırır.
    
    intervals = ['Daily', 'Hourly', 'Monthly']
    model_types = [
        'gradient_boosting',
        'random_forest',
        'decision_tree',
        'svr',
        'ridge_regression',
        'lightgbm',
        'knn'
    ]
    
    # Metrik tiplerini tanımla
    metric_types = ['r2', 'mae', 'rmse']
    
    # Her metrik tipi için ayrı sonuçlar
    all_metrics = {metric_type: {} for metric_type in metric_types}
    
    # Her zaman aralığı için metrikleri yükle
    for interval in intervals:
        # Her metrik tipi için ayrı sonuçlar
        interval_metrics = {metric_type: [] for metric_type in metric_types}
        
        # Her model tipi için metrikleri yükle
        for model_type in model_types:
            for metric_type in metric_types:
                metrics = load_performance_metrics(model_type, interval, metric_type)
                if metrics is not None:
                    interval_metrics[metric_type].append(metrics)
        
        # Her metrik tipi için sonuçları birleştir
        for metric_type in metric_types:
            if interval_metrics[metric_type]:
                # Birleştirme yaparken 'symbol' sütunu üzerinden birleştir
                if len(interval_metrics[metric_type]) > 0:
                    base_df = interval_metrics[metric_type][0]
                    if len(interval_metrics[metric_type]) > 1:
                        for df in interval_metrics[metric_type][1:]:
                            base_df = pd.merge(base_df, df, on='symbol', how='outer')
                    all_metrics[metric_type][interval] = base_df
    
    # Excel'e aktar
    export_to_excel(all_metrics['r2'], all_metrics['mae'], all_metrics['rmse'])

if __name__ == "__main__":
    display_metrics()
