import os
import joblib
import pandas as pd
import numpy as np
from tabulate import tabulate
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Karma veri seti için model performanslarını karşılaştıran ve analiz eden modül.
def load_karma_performance_metrics(model_type, interval, metric_type='r2'):
    
    # Karma veri seti için belirli bir model türü ve zaman aralığındaki performans metriklerini yükler.
    
    base_dir = "d:/Crypto_Analysis_random_state/karma_models"
    model_dir = os.path.join(base_dir, model_type, interval)
    
    if not os.path.exists(model_dir):
        # Sessiz mod - hata mesajı yazdırma
        return None
    
    # Performans dosyasını ara
    performance_file = os.path.join(model_dir, "karma_performance.joblib")
    if not os.path.exists(performance_file):
        # Sessiz mod - hata mesajı yazdırma
        return None
    
    # Performans metriklerini yükle
    performance = joblib.load(performance_file)
    
    # Metrik tipine göre sütun adını belirle
    if metric_type == 'r2':
        metric_key = 'r2_test'
        column_prefix = 'Test R²'
    elif metric_type == 'mae':
        metric_key = 'mae_test'
        column_prefix = 'Test MAE'
    elif metric_type == 'rmse':
        metric_key = 'rmse_test'
        column_prefix = 'Test RMSE'
    else:
        return None
    
    model_short_name = {
        'gradient_boosting': 'GB',
        'random_forest': 'RF',
        'decision_tree': 'DT',
        'svr': 'SVR',
        'ridge': 'Ridge',
        'lightgbm': 'LGBM',
        'knn': 'KNN'
    }.get(model_type, model_type.upper())
    
    column_name = f'{column_prefix} ({model_short_name})'
    
    # Karma modeli için tek bir satır oluştur
    metrics = [{
        'Model': 'Karma Model',
        column_name: performance.get(metric_key, 0)
    }]
    
    return pd.DataFrame(metrics)

def print_karma_model_comparison(metrics_dict, metric_type='r2'):
    # Belirli bir model türü için performans metriklerini karşılaştırır ve en iyisini belirler.
    
    # Boş metrikleri filtrele
    metrics_dict = {k: v for k, v in metrics_dict.items() if v is not None}
    
    if not metrics_dict:
        print("Hiçbir model metriği bulunamadı.")
        return
    
    # Metrik tipine göre başlık ve karşılaştırma yönünü belirle
    if metric_type == 'r2':
        metric_name = 'R²'
        is_higher_better = True
        column_prefix = 'Test R²'
    elif metric_type == 'mae':
        metric_name = 'MAE'
        is_higher_better = False
        column_prefix = 'Test MAE'
    elif metric_type == 'rmse':
        metric_name = 'RMSE'
        is_higher_better = False
        column_prefix = 'Test RMSE'
    else:
        return
    
    # Ortalama hesaplama
    avg_scores = {}
    for model_type, metrics in metrics_dict.items():
        col_name = next((col for col in metrics.columns if col.startswith(column_prefix)), None)
        if col_name:
            avg_scores[model_type] = metrics[col_name].mean()
    
    # En iyi model (R² için en yüksek, MAE/RMSE için en düşük)
    if avg_scores:
        if is_higher_better:
            winner = max(avg_scores.items(), key=lambda x: x[1])
        else:
            winner = min(avg_scores.items(), key=lambda x: x[1])
        
        # Sonuçları yazdır
        print(f"\nKarma Model {metric_name} Karşılaştırmaları:")
        print("-" * 50)
        for model_type, score in avg_scores.items():
            if metric_type == 'r2':
                print(f"{model_type} - Ortalama Test {metric_name}: {score*100:.2f}%")
            else:
                print(f"{model_type} - Ortalama Test {metric_name}: {score:.6f}")
        print("-" * 50)
        
        if metric_type == 'r2':
            print(f"En İyi Model: {winner[0]} ({metric_name} = {winner[1]*100:.2f}%)")
        else:
            print(f"En İyi Model: {winner[0]} ({metric_name} = {winner[1]:.6f})")
        print("-" * 50)

def create_karma_comparison_sheet(all_metrics, metric_type='r2'):
    # Tüm zaman aralıkları için karşılaştırma özeti oluşturur.
    
    comparison_data = []
    
    # Metrik tipine göre başlık ve karşılaştırma yönünü belirle
    if metric_type == 'r2':
        metric_name = 'R²'
        is_higher_better = True
        column_prefix = 'Test R²'
    elif metric_type == 'mae':
        metric_name = 'MAE'
        is_higher_better = False
        column_prefix = 'Test MAE'
    elif metric_type == 'rmse':
        metric_name = 'RMSE'
        is_higher_better = False
        column_prefix = 'Test RMSE'
    else:
        return pd.DataFrame()
    
    # Her zaman aralığı için performansları ekle
    for interval, metrics_dict in all_metrics.items():
        row_data = {'Zaman Aralığı': interval.capitalize()}
        
        # Her model türü için metrik değerini ekle
        for model_type, metrics in metrics_dict.items():
            if metrics is not None:
                col_name = next((col for col in metrics.columns if col.startswith(column_prefix)), None)
                if col_name:
                    model_short_name = col_name.split('(')[1].split(')')[0]
                    row_data[f'{model_short_name} {metric_name}'] = metrics[col_name].mean()
        
        # En iyi model ve skoru
        model_scores = {k.split(' ')[0]: v for k, v in row_data.items() if f' {metric_name}' in k}
        if model_scores:
            if is_higher_better:
                best_model = max(model_scores.items(), key=lambda x: x[1])
            else:
                best_model = min(model_scores.items(), key=lambda x: x[1])
                
            row_data['En İyi Model'] = best_model[0]
            row_data['En İyi Skor'] = best_model[1]
        
        comparison_data.append(row_data)
    
    # DataFrame oluştur
    if not comparison_data:
        return pd.DataFrame()
    
    comparison_df = pd.DataFrame(comparison_data)
    
    # Sütun sıralamasını düzenle
    all_columns = list(comparison_df.columns)
    ordered_columns = ['Zaman Aralığı', 'En İyi Model', 'En İyi Skor']
    other_columns = [col for col in all_columns if col not in ordered_columns]
    column_order = ordered_columns + sorted(other_columns)
    
    # Sadece var olan sütunları seç
    column_order = [col for col in column_order if col in all_columns]
    comparison_df = comparison_df[column_order]
    
    # Sayısal değerleri formatla
    for col in comparison_df.columns:
        if col not in ['Zaman Aralığı', 'En İyi Model']:
            if metric_type == 'r2':
                comparison_df[col] = comparison_df[col].apply(lambda x: f"{x*100:.2f}%")
            else:
                comparison_df[col] = comparison_df[col].apply(lambda x: f"{x:.6f}")
    
    return comparison_df

def apply_excel_styling(worksheet, df, metric_type='r2'):
   # Excel sayfasına stil uygular.
    
    # Renk tanımları
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    
    # Metrik tipine göre en iyi skor rengi belirle
    if metric_type == 'r2':
        best_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Yeşil (yüksek daha iyi)
    else:
        best_score_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Kırmızı (düşük daha iyi)
    
    # Kenarlık stilleri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlık stili
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    
    # Başlıkları formatla
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Verileri formatla
    for row_num, row in enumerate(df.itertuples(index=False), 2):
        # Alternatif satır renklendirmesi
        row_fill = alt_row_fill if row_num % 2 == 0 else None
        
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            if row_fill:
                cell.fill = row_fill
            
            # En iyi skoru vurgula
            if df.columns[col_num-1] == 'En İyi Skor':
                cell.fill = best_score_fill
    
    # Sütun genişliklerini ayarla
    for col_num, column in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        column_width = max(len(str(column)), df[column].astype(str).map(len).max())
        worksheet.column_dimensions[column_letter].width = column_width + 4

def export_to_excel(all_metrics_r2, all_metrics_mae, all_metrics_rmse):
    # Tüm metrikleri Excel dosyasına aktarır.
   
    # Karşılaştırma sayfaları oluştur
    comparison_df_r2 = create_karma_comparison_sheet(all_metrics_r2, 'r2')
    comparison_df_mae = create_karma_comparison_sheet(all_metrics_mae, 'mae')
    comparison_df_rmse = create_karma_comparison_sheet(all_metrics_rmse, 'rmse')
    
    # Excel dosyası oluştur
    output_dir = "d:/Crypto_Analysis_random_state/sonuçlar"
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, "karma_sonuclar.xlsx")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Karşılaştırma sayfaları
        comparison_df_r2.to_excel(writer, sheet_name='R² Karşılaştırma', index=False)
        apply_excel_styling(writer.sheets['R² Karşılaştırma'], comparison_df_r2, 'r2')
        
        comparison_df_mae.to_excel(writer, sheet_name='MAE Karşılaştırma', index=False)
        apply_excel_styling(writer.sheets['MAE Karşılaştırma'], comparison_df_mae, 'mae')
        
        comparison_df_rmse.to_excel(writer, sheet_name='RMSE Karşılaştırma', index=False)
        apply_excel_styling(writer.sheets['RMSE Karşılaştırma'], comparison_df_rmse, 'rmse')
    
    print(f"\nSonuçlar başarıyla kaydedildi: {output_file}")

def display_karma_metrics():
    # Karma model metriklerini görüntüler ve Excel'e aktarır.
    print("\nKarma Model Performans Analizi")
    print("=" * 50)
    
    intervals = ['daily', 'hourly', 'monthly']
    model_types = ['dtr', 'rf', 'gradient_boosting', 'svr', 'ridge', 'lightgbm', 'knn']
    
    # Tüm metrikleri topla
    all_metrics_r2 = {}
    all_metrics_mae = {}
    all_metrics_rmse = {}
    
    for interval in intervals:
        # Her metrik tipi için ayrı sonuçlar
        interval_metrics_r2 = {}
        interval_metrics_mae = {}
        interval_metrics_rmse = {}
        
        for model_type in model_types:
            # R² metrikleri
            metrics_r2 = load_karma_performance_metrics(model_type, interval, 'r2')
            if metrics_r2 is not None:
                interval_metrics_r2[model_type] = metrics_r2
            
            # MAE metrikleri
            metrics_mae = load_karma_performance_metrics(model_type, interval, 'mae')
            if metrics_mae is not None:
                interval_metrics_mae[model_type] = metrics_mae
            
            # RMSE metrikleri
            metrics_rmse = load_karma_performance_metrics(model_type, interval, 'rmse')
            if metrics_rmse is not None:
                interval_metrics_rmse[model_type] = metrics_rmse
        
        # R² karşılaştırmaları
        if interval_metrics_r2:
            all_metrics_r2[interval] = interval_metrics_r2
            print_karma_model_comparison(interval_metrics_r2, 'r2')
        
        # MAE karşılaştırmaları
        if interval_metrics_mae:
            all_metrics_mae[interval] = interval_metrics_mae
            print_karma_model_comparison(interval_metrics_mae, 'mae')
        
        # RMSE karşılaştırmaları
        if interval_metrics_rmse:
            all_metrics_rmse[interval] = interval_metrics_rmse
            print_karma_model_comparison(interval_metrics_rmse, 'rmse')
    
    if all_metrics_r2 or all_metrics_mae or all_metrics_rmse:
        export_to_excel(all_metrics_r2, all_metrics_mae, all_metrics_rmse)
    else:
        print("\nHiçbir karma model metriği bulunamadı.")

if __name__ == "__main__":
    display_karma_metrics()
